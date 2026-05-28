from pathlib import Path
import json
import re
import pandas as pd
import openpyxl

from metric_catalog import load_metric_catalog


UPLOAD_DIR = Path("uploads")
REPOSITORY_DIR = Path("repository")


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_numeric(value):
    return isinstance(value, (int, float)) and not pd.isna(value)


def normalize_text(value):
    return clean_text(value).lower()


def cell_address(row, col):
    return openpyxl.utils.get_column_letter(col) + str(row)


# Column-header keywords used by table-aware extraction. Order = preference.
# When the header row of a table contains one of these keywords, the matching
# column is preferred over generic "first value to the right".
_TOTAL_COLUMN_KEYWORDS    = ["total", "total cost", "all-in", "aggregate", "lifetime", "cumulative"]
_PERIOD_COLUMN_KEYWORDS   = ["at close", "closing", "initial", "year 1", "y1", "yr 1",
                             "post close", "post-close", "stabilized", "stable", "exit", "year 5", "yr 5"]
_PER_UNIT_COLUMN_KEYWORDS = ["$/unit", "$/sf", "$/gsf", "$/nsf", "per unit", "per sf", "% total", "% of total"]


def _scan_data_columns(ws, label_row, label_col, max_scan=10):
    """Return list of column indices to the right of label_col that contain numeric values."""
    data_cols = []
    for offset in range(1, max_scan + 1):
        cell = ws.cell(row=label_row, column=label_col + offset)
        if is_numeric(cell.value):
            data_cols.append(label_col + offset)
        elif cell.value is not None and str(cell.value).strip():
            # Hit a text cell — table row ends here
            break
    return data_cols


def _detect_column_headers(ws, label_row, data_cols, max_lookback=10):
    """
    Look up from the labeled row to find the nearest header row.

    A row qualifies as a header if:
      - At least half of the data_cols have non-numeric text values in that row
      - Headers are typical column descriptors (years, periods, "Total", etc.)

    Returns {col_idx: header_text}.
    """
    if not data_cols:
        return {}

    for r_offset in range(1, max_lookback + 1):
        header_row = label_row - r_offset
        if header_row < 1:
            break
        headers = {}
        for col in data_cols:
            val = ws.cell(row=header_row, column=col).value
            if isinstance(val, str) and val.strip() and not is_numeric(val):
                headers[col] = val.strip()
        # Accept this as a header row if most data cols are text-headed
        if len(headers) >= max(2, len(data_cols) // 2):
            return headers
    return {}


def _pick_column_for_metric(headers: dict, metric_name_lower: str) -> int | None:
    """
    Given a {col: header_text} map and a metric name, pick the best column.

    Logic (in order):
      1. If metric name mentions a specific period (e.g. "Year 1", "At Close",
         "Stabilized"), use the column whose header matches that period.
      2. If metric name implies a total ("Total X", "All-in", "Project Cost"),
         use the column whose header matches Total-like keywords.
      3. Skip per-unit / per-SF / % columns (those are derived metrics, not the value).
      4. Otherwise return None — caller falls back to "first value" behavior.
    """
    headers_lower = {col: h.lower() for col, h in headers.items()}

    # (1) Period-specific preference based on metric name
    period_map = [
        (["at close", "closing", "initial", "going-in", "going in", "purchase"],
            ["at close", "closing", "initial", "going-in", "going in"]),
        (["post close", "post-close", "draws", "construction"],
            ["post close", "post-close", "draws", "construction"]),
        (["stabilized", "stabilization", "stable"],
            ["stabilized", "stable", "stab"]),
        (["year 1", "y1", "yr 1", "first year"],
            ["year 1", "y1", "yr 1"]),
        (["exit", "year 5", "yr 5", "terminal", "disposition"],
            ["exit", "year 5", "yr 5", "terminal"]),
    ]
    for metric_keywords, header_keywords in period_map:
        if any(mk in metric_name_lower for mk in metric_keywords):
            for col, h in headers_lower.items():
                if any(hk in h for hk in header_keywords):
                    return col

    # (2) Total-like preference (default for cost/proceeds/sources/uses items)
    # If ANY column header is "Total" or similar, prefer it.
    for col, h in headers_lower.items():
        if any(kw == h or kw in h for kw in _TOTAL_COLUMN_KEYWORDS):
            # But skip if it's actually a per-unit column ("$/Total" doesn't exist
            # but be safe)
            if not any(pu in h for pu in _PER_UNIT_COLUMN_KEYWORDS):
                return col

    return None


def find_nearby_value(ws, row, col, metric_name: str = ""):
    """
    Find the value associated with a labeled cell.

    Strategy:
      1. Scan data columns to the right
      2. If multiple columns exist, look UP for a header row
      3. If headers found, pick the column best matching the metric's semantics
         (Total column for total metrics, period-specific for period metrics)
      4. Otherwise fall back to first non-zero value
      5. If no values right, look below
      6. Last resort: nearby grid scan
    """

    data_cols = _scan_data_columns(ws, row, col)

    # Table-aware path: 2+ columns of data → likely a table
    if len(data_cols) >= 2:
        headers = _detect_column_headers(ws, row, data_cols)
        if headers and metric_name:
            preferred_col = _pick_column_for_metric(headers, metric_name.lower())
            if preferred_col is not None:
                value = ws.cell(row=row, column=preferred_col).value
                if is_numeric(value):
                    return value, cell_address(row, preferred_col), "table-column"

        # Skip per-unit/percentage columns when picking fallback
        non_derived_cols = [
            c for c in data_cols
            if not (headers and any(
                pu in headers.get(c, "").lower() for pu in _PER_UNIT_COLUMN_KEYWORDS
            ))
        ] or data_cols

        right_values = [
            (ws.cell(row=row, column=c).value, cell_address(row, c))
            for c in non_derived_cols
        ]
        non_zero = [(v, a) for v, a in right_values if v != 0]
        best_val, best_addr = (non_zero[0] if non_zero else right_values[0])
        return best_val, best_addr, "right"

    # Single value to the right (no table) — return it
    if len(data_cols) == 1:
        c = data_cols[0]
        return ws.cell(row=row, column=c).value, cell_address(row, c), "right"

    # Look below
    for offset in range(1, 6):
        value = ws.cell(row=row + offset, column=col).value
        if is_numeric(value):
            return value, cell_address(row + offset, col), "below"

    # Last resort: nearby grid scan
    for r_offset in range(-2, 4):
        for c_offset in range(-2, 6):
            r = row + r_offset
            c = col + c_offset
            if r < 1 or c < 1:
                continue
            value = ws.cell(row=r, column=c).value
            if is_numeric(value):
                return value, cell_address(r, c), "nearby"

    return None, None, None


# Which data_nature values are relevant per SSOT layer.
# "mixed" metrics are always included (meaningful in both projection and actual contexts).
# "underwriting" scans all three because acquisition/closing models routinely contain both
# projected values (IRR, NOI proforma, exit cap) and actual values (closing costs paid,
# loan amount drawn, actual purchase price confirmed at closing).
_LAYER_DATA_NATURE: dict[str, set] = {
    "underwriting":    {"projection", "actual", "mixed"},
    "business_plan":   {"projection", "actual", "mixed"},
    "actuals_2020":    {"actual", "mixed"},
    "actuals_2021":    {"actual", "mixed"},
    "actuals_2022":    {"actual", "mixed"},
    "actuals_2023":    {"actual", "mixed"},
    "actuals_2024":    {"actual", "mixed"},
    "actuals_2025":    {"actual", "mixed"},
    "actuals_recent":  {"actual", "mixed"},
    "rent_roll":       {"actual", "mixed"},
    "debt":            {"actual", "mixed"},
}


def filter_catalog_for_layer(catalog: list, layer: str) -> list:
    """
    Return only the metrics relevant to a given SSOT layer.

    Two filters applied:
    1. Skip calculated metrics (metric_source == "calculated") — these are
       derived after extraction, not extracted from cells.
    2. Keep only metrics whose data_nature matches the layer's expected type.
       e.g. an underwriting file should not be scanned for Current LTV or DSCR
       (those are actual/current-state metrics).
    """
    allowed_natures = _LAYER_DATA_NATURE.get(layer, {"projection", "actual", "mixed"})
    return [
        m for m in catalog
        if m.get("metric_source", "extracted") == "extracted"
        and m.get("data_nature", "mixed") in allowed_natures
    ]


def scan_workbook_for_all_metrics(file_path, catalog):
    """
    Load the workbook ONCE and scan all catalog metrics in a single pass.

    This is the fast path used by v2's tools.extract_from_file. It replaces
    the prior pattern of calling scan_workbook_for_metric in a loop, which
    re-loaded the same Excel file once per metric (≈97x per file).

    Returns {metric_id: best_match_dict_or_None} for every metric in the catalog.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return {m["metric_id"]: None for m in catalog}

    # Pre-normalize every alias once, paired with its parent metric.
    # Each entry: (normalized_alias_text, metric_dict, original_alias_string)
    alias_index = []
    for metric in catalog:
        for alias in metric.get("aliases", []):
            alias_text = normalize_text(alias)
            if alias_text:
                alias_index.append((alias_text, metric, alias))

    matches_by_metric: dict = {m["metric_id"]: [] for m in catalog}
    file_name = Path(file_path).name

    # Scan high-signal sheets first so their matches win ties.
    # Annual summary/assumption sheets contain the definitive values;
    # monthly detail tabs often have the same labels with partial/zero values.
    _PRIORITY_SHEET_KEYWORDS = [
        "summary", "assumption", "sources", "uses", "return",
        "waterfall", "annual", "overview", "irr", "exit",
        "proforma", "pro forma", "debt", "equity",
    ]

    def _sheet_priority(name: str) -> int:
        nl = name.lower()
        if any(kw in nl for kw in _PRIORITY_SHEET_KEYWORDS):
            return 0
        if "monthly" in nl or "month" in nl:
            return 2   # monthly detail — scan last
        return 1

    sorted_sheet_names = sorted(wb.sheetnames, key=_sheet_priority)

    for sheet_name in sorted_sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell_text = normalize_text(cell.value)
                if not cell_text:
                    continue

                for alias_text, metric, original_alias in alias_index:
                    if alias_text not in cell_text:
                        continue

                    # Label quality: penalise matches where the alias is a small
                    # fraction of the cell label (e.g. "noi" inside "noi to offset
                    # interest"). An exact or near-exact label match scores 1.0;
                    # a substring-in-long-label scores proportionally lower.
                    label_ratio = len(alias_text) / max(len(cell_text), 1)
                    # Also penalise if alias appears mid-word (e.g. "irr" in "irrespective")
                    idx = cell_text.find(alias_text)
                    char_before = cell_text[idx - 1] if idx > 0 else " "
                    char_after  = cell_text[idx + len(alias_text)] if idx + len(alias_text) < len(cell_text) else " "
                    mid_word = char_before.isalpha() or char_after.isalpha()
                    if mid_word:
                        continue  # alias embedded inside a longer word — skip

                    value, value_cell, direction = find_nearby_value(
                        ws, cell.row, cell.column,
                        metric_name=metric["metric_name"],
                    )
                    if value is None:
                        continue

                    # Confidence tiers:
                    #   "exact"  — alias covers ≥80% of the cell label, value right/below
                    #   "high"   — value right/below (alias may be partial label)
                    #   "medium" — value found nearby
                    #   "partial"— alias is a small fragment of a longer label (label_ratio < 0.4)
                    if direction in ("right", "below"):
                        confidence = "exact" if label_ratio >= 0.8 else "high"
                    else:
                        confidence = "partial" if label_ratio < 0.4 else "medium"

                    matches_by_metric[metric["metric_id"]].append({
                        "metric_id": metric["metric_id"],
                        "metric_name": metric["metric_name"],
                        "category": metric["category"],
                        "definition": metric["definition"],
                        "value": value,
                        "source_file": file_name,
                        "sheet": sheet_name,
                        "label_cell": cell.coordinate,
                        "value_cell": value_cell,
                        "matched_alias": original_alias,
                        "confidence": confidence,
                        "label_ratio": round(label_ratio, 2),
                        "match_method": direction,
                    })

    # Best match per metric — ranked by confidence tier then label quality.
    # Tier order: exact > high > medium > partial
    _TIER = {"exact": 0, "high": 1, "medium": 2, "partial": 3}
    best = {}
    for metric_id, matches in matches_by_metric.items():
        if not matches:
            best[metric_id] = None
        else:
            matches.sort(key=lambda x: (
                _TIER.get(x["confidence"], 9),
                -x.get("label_ratio", 0),  # higher label_ratio wins ties
            ))
            best[metric_id] = matches[0]
    return best


def scan_workbook_for_metric(file_path, metric):
    """
    Search one Excel workbook for one metric.
    Returns best match or None.

    NOTE: kept for backward compatibility with v1 modules. The fast path is
    scan_workbook_for_all_metrics, which avoids reloading the workbook per metric.
    """

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None

    aliases = metric.get("aliases", [])
    matches = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                cell_text = normalize_text(cell.value)

                if not cell_text:
                    continue

                for alias in aliases:
                    alias_text = normalize_text(alias)

                    if not alias_text:
                        continue

                    if alias_text in cell_text:
                        value, value_cell, direction = find_nearby_value(
                            ws,
                            cell.row,
                            cell.column,
                            metric_name=metric["metric_name"],
                        )

                        if value is not None:
                            confidence = "high" if direction in ["right", "below"] else "medium"

                            matches.append({
                                "metric_id": metric["metric_id"],
                                "metric_name": metric["metric_name"],
                                "category": metric["category"],
                                "definition": metric["definition"],
                                "value": value,
                                "source_file": Path(file_path).name,
                                "sheet": sheet_name,
                                "label_cell": cell.coordinate,
                                "value_cell": value_cell,
                                "matched_alias": alias,
                                "confidence": confidence,
                                "match_method": direction,
                            })

    if not matches:
        return None

    # Prefer high confidence matches first
    matches = sorted(
        matches,
        key=lambda x: 0 if x["confidence"] == "high" else 1
    )

    return matches[0]


def extract_raw_labeled_pairs(file_path, max_pairs: int = 600) -> list[dict]:
    """
    Extract ALL (sheet, label, value) pairs from a workbook without any
    catalog filtering. This is the input for Pass 2 (GPT insight pass).

    Returns a list of dicts:
        {"sheet": str, "label": str, "value": numeric, "cell": str,
         "direction": "right"|"below"|"nearby", "label_len": int}

    Quality fields (used by run_raw_insight_pass to filter noise):
      direction: "right"/"below" = label directly precedes value — high signal
                 "nearby" = value found in surrounding area — lower signal
      label_len: very short labels (< 5 chars) are often headers/indices, not metrics

    Capped at max_pairs. Priority sheets (summary, assumptions, waterfall) come first.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    # Prioritise summary/assumption sheets so if we hit the cap we keep the
    # most analytically relevant rows.
    priority_keywords = [
        "summary", "assumption", "return", "waterfall", "overview",
        "sources", "uses", "debt", "equity", "cashflow", "cash flow",
        "proforma", "pro forma", "irr", "exit",
    ]

    def sheet_priority(name: str) -> int:
        nl = name.lower()
        return 0 if any(kw in nl for kw in priority_keywords) else 1

    sorted_sheets = sorted(wb.sheetnames, key=sheet_priority)

    pairs = []
    seen_labels: set[str] = set()

    for sheet_name in sorted_sheets:
        if len(pairs) >= max_pairs:
            break
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            if len(pairs) >= max_pairs:
                break
            for cell in row:
                cell_text = clean_text(cell.value)
                if not cell_text or len(cell_text) < 3:
                    continue
                # Only process text cells (labels)
                if not isinstance(cell.value, str):
                    continue

                value, value_cell, direction = find_nearby_value(
                    ws, cell.row, cell.column
                )
                if value is None:
                    continue

                # Deduplicate by (sheet, normalised label) to avoid
                # repeated header rows skewing GPT's reading.
                key = f"{sheet_name}|{normalize_text(cell_text)}"
                if key in seen_labels:
                    continue
                seen_labels.add(key)

                pairs.append({
                    "sheet":     sheet_name,
                    "label":     cell_text,
                    "value":     value,
                    "cell":      value_cell,
                    "direction": direction,
                    "label_len": len(cell_text),
                })

    return pairs


def extract_time_series_rows(file_path, max_rows_per_sheet: int = 25, max_total_rows: int = 80) -> list[dict]:
    """
    Find rows in the workbook that look like multi-year time series.

    A row qualifies as a time series if:
      - It has a text label in the leftmost data column
      - 3+ numeric cells follow in consecutive columns
      - A row above has text headers that look like years (2020-2035) or
        period labels (Y1, Yr 1, Year 1, Q1, Stabilized, Exit, etc.)

    Returns list of:
      {
        "sheet": str,
        "label": str,             # row label (e.g. "Net Operating Income")
        "label_cell": str,        # cell ref of label
        "headers": [str],         # column headers (years/periods)
        "values": [number],       # aligned with headers
      }

    Cap at max_rows to avoid huge payloads. Prioritises sheets named like
    cash flow projections.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    # Prioritise cash flow / proforma sheets
    cf_keywords = ["annual", "cash flow", "cashflow", "cf", "proforma", "pro forma",
                   "operating", "monthly", "schedule"]

    def sheet_priority(name: str) -> int:
        nl = name.lower()
        return 0 if any(kw in nl for kw in cf_keywords) else 1

    sorted_sheets = sorted(wb.sheetnames, key=sheet_priority)

    # Labels to skip — meta/structural rows that aren't analytically interesting
    _NOISE_LABELS = {"year", "month", "period", "day", "date", "row", "n/a"}

    series = []
    series_per_sheet: dict[str, int] = {}
    period_pattern_re = re.compile(
        r"^(20\d{2}|y(ear)?\s*\d{1,2}|yr\s*\d{1,2}|q[1-4]|fy\d{2,4}|"
        r"stabili[sz]ed|exit|going.?in|at.close|post.close|trended|untrended)",
        re.IGNORECASE,
    )

    def looks_like_period_header(val) -> bool:
        """Accepts strings matching the regex OR integers in year range (2000-2100)."""
        if val is None:
            return False
        # Year stored as a number
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return 2000 <= val <= 2100
        s = str(val).strip()
        if not s:
            return False
        return bool(period_pattern_re.match(s))

    for sheet_name in sorted_sheets:
        if len(series) >= max_total_rows:
            break
        ws = wb[sheet_name]
        # Pre-scan: find candidate header rows (rows where most cells are period-like)
        header_rows: list[tuple[int, dict[int, str]]] = []
        for r in range(1, min(ws.max_row, 200)):
            row_headers: dict[int, str] = {}
            period_count = 0
            for c in range(1, min(ws.max_column, 30) + 1):
                v = ws.cell(row=r, column=c).value
                if looks_like_period_header(v):
                    row_headers[c] = str(v).strip() if not isinstance(v, (int, float)) else str(int(v))
                    period_count += 1
            if period_count >= 3:
                header_rows.append((r, row_headers))

        if not header_rows:
            continue

        # For each header row, look at subsequent rows for label + values matching the header columns
        for header_row, headers in header_rows:
            header_cols = sorted(headers.keys())
            for r in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
                if len(series) >= max_total_rows:
                    break
                if series_per_sheet.get(sheet_name, 0) >= max_rows_per_sheet:
                    break

                # Look left of the first header column for a text label
                label = None
                label_cell = None
                for c in range(1, header_cols[0]):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, str) and v.strip() and len(v.strip()) >= 3:
                        label = v.strip()
                        label_cell = cell_address(r, c)
                        break
                if not label:
                    continue

                # Skip noise labels (Year, Month, etc. — not analytically meaningful)
                if label.lower().strip(":") in _NOISE_LABELS:
                    continue

                # Collect values aligned with header columns
                values = []
                aligned_headers = []
                for c in header_cols:
                    v = ws.cell(row=r, column=c).value
                    if is_numeric(v):
                        values.append(v)
                        aligned_headers.append(headers[c])
                    else:
                        values.append(None)
                        aligned_headers.append(headers[c])

                numeric_count = sum(1 for v in values if v is not None)
                non_zero_count = sum(1 for v in values if v is not None and v != 0)

                # Skip rows where all values are zero (no signal — typically empty
                # construction draw rows in dev models)
                if non_zero_count == 0:
                    continue

                if numeric_count >= 3:
                    series.append({
                        "sheet":      sheet_name,
                        "label":      label,
                        "label_cell": label_cell,
                        "headers":    aligned_headers,
                        "values":     values,
                    })
                    series_per_sheet[sheet_name] = series_per_sheet.get(sheet_name, 0) + 1

    return series


def classify_file_layer(file_name):
    """
    Classify a file by its investment lifecycle layer based on its name.
    Returns one of: 'underwriting', 'business_plan', 'actuals_2021',
    'actuals_2022', 'actuals_recent', or 'unknown'.

    These names must match ssot.KNOWN_LAYERS exactly.

    Keyword groups reflect institutional RE naming conventions:
      - 'proforma' / 'pro forma' is the most common name for an UW model
      - 'BP' alone is risky (matches too much) so we anchor with word boundaries
      - financial statements: 'fs', 'financial', 'p&l', 'income statement',
        'operating statement', 't12'
    """
    name_lower = file_name.lower()

    # --- Financial Statements / actuals (check first; "2022 P&L" should NOT
    # match business plan via the year). ---
    # We pad with leading/trailing spaces so " fs " matches "FS 2022.xlsx"
    padded = f" {name_lower} "
    actuals_keywords = [
        "financial statement", "income statement", "operating statement",
        "p&l", "pl statement", "actual", "actuals",
        " fs ", "_fs_", "_fs.", " fs.", "t12", "trailing 12",
    ]
    if any(kw in padded for kw in actuals_keywords):
        for year in ("2020", "2021", "2022", "2023", "2024", "2025"):
            if year in name_lower:
                return f"actuals_{year}"
        return "actuals_recent"

    # --- Acquisition Underwriting (proforma / UW model / deal memo / closing docs) ---
    uw_keywords = [
        "acquisition", "underwriting",
        "proforma", "pro forma", "pro-forma",
        "uw model", "deal memo",
        "closing", "settlement",  # closing statement / settlement statement
        "psa", "purchase agreement",        # purchase & sale agreement
        "ic memo", "investment committee",  # IC package
    ]
    # Word-boundary check for the short token " uw" (avoid matching "answer"!)
    uw_token_match = (
        " uw" in name_lower or "_uw" in name_lower
        or name_lower.endswith(" uw") or name_lower.endswith("_uw")
    )
    if any(kw in name_lower for kw in uw_keywords) or uw_token_match:
        return "underwriting"

    # --- Business Plan (revised plan post-acquisition) ---
    bp_keywords = [
        "business plan", "budget", "forecast", "revised plan",
        "annual plan", "asset plan", "hold plan",
    ]
    if any(kw in name_lower for kw in bp_keywords):
        return "business_plan"
    # " bp " as a standalone token (so "abp_2022.xlsx" doesn't false-match)
    if " bp " in name_lower or "_bp_" in name_lower or "_bp." in name_lower or " bp." in name_lower:
        return "business_plan"

    return "unknown"


def scan_uploaded_files(upload_dir=UPLOAD_DIR):
    """
    Scan all uploaded Excel files against the metric catalog.
    Extracts each metric from EVERY file where found, tagged by source layer,
    so the analysis can compare underwriting vs business plan vs actuals.
    """

    upload_dir = Path(upload_dir)
    REPOSITORY_DIR.mkdir(exist_ok=True)

    catalog = load_metric_catalog()

    excel_files = list(upload_dir.glob("*.xlsx")) + list(upload_dir.glob("*.xlsm"))

    extracted = []
    missing = []

    for metric in catalog:
        all_matches = []

        for file_path in excel_files:
            match = scan_workbook_for_metric(file_path, metric)

            if match:
                match["source_layer"] = classify_file_layer(file_path.name)
                all_matches.append(match)

        if all_matches:
            extracted.extend(all_matches)
        else:
            missing.append({
                "metric_id": metric["metric_id"],
                "metric_name": metric["metric_name"],
                "category": metric["category"],
                "definition": metric["definition"],
                "source": metric.get("source", ""),
                "priority": metric.get("priority", "medium"),
                "aliases": metric.get("aliases", []),
                "status": "missing"
            })

    result = {
        "status": "success",
        "total_metrics": len(catalog),
        "extracted_count": len(extracted),
        "missing_count": len(missing),
        "extracted_metrics": extracted,
        "missing_metrics": missing,
    }

    with open(REPOSITORY_DIR / "flexible_extraction_result.json", "w") as f:
        json.dump(result, f, indent=2, default=str)

    pd.DataFrame(extracted).to_csv(
        REPOSITORY_DIR / "extracted_metrics_report.csv",
        index=False
    )

    pd.DataFrame(missing).to_csv(
        REPOSITORY_DIR / "missing_metrics_report.csv",
        index=False
    )

    return result


if __name__ == "__main__":
    result = scan_uploaded_files()
    print(f"Total metrics: {result['total_metrics']}")
    print(f"Extracted: {result['extracted_count']}")
    print(f"Missing: {result['missing_count']}")
    print("Saved reports to repository/")