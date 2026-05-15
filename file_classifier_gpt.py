import os
import json
from pathlib import Path

import openpyxl
import streamlit as st
from openai import OpenAI


try:
    api_key = st.secrets["OPENAI_API_KEY"]
except Exception:
    api_key = os.getenv("OPENAI_API_KEY")


# Defer client creation so missing keys raise at call time (not import time)
client = OpenAI(api_key=api_key) if api_key else None


UPLOAD_DIR = Path("uploads")
REPOSITORY_DIR = Path("repository")


CLASSIFIER_SYSTEM_PROMPT = """
You are a real estate investment management file classifier.

Your job is to review lightweight previews of uploaded files and classify each file by its likely role in an institutional real estate asset management workflow.

You are not analyzing investment performance yet.
You are only classifying what each file appears to be and where it sits in the investment lifecycle.

Common document types:
- acquisition_underwriting
- business_plan
- financial_statement_actuals
- rent_roll
- debt_model
- capex_tracker
- lease_abstract
- investor_reporting
- valuation_model
- waterfall_model
- market_research
- unknown_or_unsupported

Investment lifecycle roles:
- original_investment_thesis
- updated_operating_forecast
- realized_operating_performance
- leasing_income_durability
- leverage_capital_structure
- capital_project_execution
- valuation_return_analysis
- investor_lender_reporting
- unknown

Return only valid JSON. Do not include markdown.
"""


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def preview_excel_file(file_path, max_sheets=8, max_rows=35, max_cols=12):
    """
    Creates a lightweight preview of an Excel workbook:
    - filename
    - sheet names
    - first rows/columns from each sheet
    """

    preview = {
        "file_name": Path(file_path).name,
        "file_extension": Path(file_path).suffix,
        "sheets": [],
    }

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        preview["error"] = f"Could not open workbook: {str(e)}"
        return preview

    for sheet_name in wb.sheetnames[:max_sheets]:
        ws = wb[sheet_name]

        sheet_preview = {
            "sheet_name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample_rows": [],
        }

        for row in ws.iter_rows(
            min_row=1,
            max_row=min(max_rows, ws.max_row),
            min_col=1,
            max_col=min(max_cols, ws.max_column),
            values_only=True
        ):
            cleaned_row = [clean_value(cell) for cell in row]
            if any(cleaned_row):
                sheet_preview["sample_rows"].append(cleaned_row)

        preview["sheets"].append(sheet_preview)

    return preview


def preview_uploaded_files(upload_dir=UPLOAD_DIR):
    upload_dir = Path(upload_dir)

    previews = []

    files = list(upload_dir.glob("*"))

    for file_path in files:
        if file_path.name.startswith("."):
            continue

        suffix = file_path.suffix.lower()

        if suffix in [".xlsx", ".xlsm"]:
            previews.append(preview_excel_file(file_path))
        elif suffix == ".csv":
            previews.append({
                "file_name": file_path.name,
                "file_extension": suffix,
                "note": "CSV preview not implemented yet."
            })
        elif suffix == ".pdf":
            previews.append({
                "file_name": file_path.name,
                "file_extension": suffix,
                "note": "PDF preview not implemented yet."
            })
        else:
            previews.append({
                "file_name": file_path.name,
                "file_extension": suffix,
                "note": "Unsupported file type for preview."
            })

    return previews


def classify_file_preview(file_preview):
    prompt = {
        "task": "Classify this uploaded file based on the preview.",
        "file_preview": file_preview,
        "required_json_schema": {
            "file_name": "string",
            "document_type": "one of the common document types",
            "investment_lifecycle_role": "one of the lifecycle roles",
            "likely_period": "string or null",
            "likely_year": "integer or null",
            "relevant_tabs": ["list of relevant worksheet names"],
            "key_detected_sections": ["list of important detected labels or sections"],
            "recommended_extraction_modules": ["list of recommended extraction approaches"],
            "confidence": "high | medium | low",
            "reasoning": "brief explanation"
        }
    }

    response = client.responses.create(
        model="gpt-5.5",
        input=[
            {"role": "system", "content": CLASSIFIER_SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(prompt, default=str)}
        ],
    )

    text = response.output_text.strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {
            "file_name": file_preview.get("file_name"),
            "document_type": "unknown_or_unsupported",
            "investment_lifecycle_role": "unknown",
            "likely_period": None,
            "likely_year": None,
            "relevant_tabs": [],
            "key_detected_sections": [],
            "recommended_extraction_modules": [],
            "confidence": "low",
            "reasoning": "Classifier response was not valid JSON.",
            "raw_response": text,
        }


def classify_uploaded_files(upload_dir=UPLOAD_DIR):
    REPOSITORY_DIR.mkdir(exist_ok=True)

    previews = preview_uploaded_files(upload_dir)

    classifications = []

    for preview in previews:
        classification = classify_file_preview(preview)
        classifications.append(classification)

    result = {
        "status": "success",
        "file_count": len(classifications),
        "classifications": classifications,
        "previews": previews,
    }

    with open(REPOSITORY_DIR / "file_classification_result.json", "w") as f:
        json.dump(result, f, indent=2, default=str)

    return result


if __name__ == "__main__":
    result = classify_uploaded_files()
    print(json.dumps(result["classifications"], indent=2))
