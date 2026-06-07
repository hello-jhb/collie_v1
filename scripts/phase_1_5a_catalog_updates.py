"""
phase_1_5a_catalog_updates.py — Catalog/alias/schema tuning for Phase 1.5a.

What this does:
  1. Adds new bounded metrics:
     - Number of Properties      (count, multi-property detection)
     - Purchase Date             (date, for Hold Period derivation)
     - Exit Date                 (date, for Hold Period derivation)
     - Interest Rate Spread      (ratio, floating-rate component)
     - Interest Rate Cap         (ratio, floating-rate cap strike)
  2. Expands aliases on existing metrics that BAC test exposed as too narrow:
     - Physical Occupancy        (add "Occ %", "% Occupied", "Occupancy", etc.)
     - Interest-Only Period      (add "I/O Period", "IO Months", etc.)
     - Interest Rate             (add "All-in Rate", "Coupon", "Pay Rate")
     - Hold Period               (tighten — remove "Hold" alone which matches wrong cells)
     - CapEx Budget              (add "Total CapEx", "Renovation Cost", etc.)
  3. Adds the `applies_to_property_types` column:
     - Empty / blank → applies to all
     - Comma list    → only applies if asset's property_type matches
     - Used to skip Total SF for Hotels (hotels track Keys, not SF)
"""
from __future__ import annotations
from openpyxl import load_workbook

CATALOG_PATH = "Snapshot Metric.xlsx"


# ----------------------------------------------------------------------------
# NEW METRICS — appended to catalog
# ----------------------------------------------------------------------------
NEW_METRICS = [
    {
        "metric_name": "Number of Properties",
        "metric_source": "extracted",
        "category": "Investment Basis",
        "definition": "Number of distinct assets / properties in this deal (1 for single-asset, 2+ for portfolio)",
        "source_type": "General Information, Investment Summary",
        "data_nature": "mixed",
        "priority": "High",
        "core_question": "Is this a single-asset or portfolio deal?",
        "aliases": "Number of Properties; Number of Properties/Assets; # of Properties; # of Assets; Number of Hotels; # of Hotels; Number of Buildings; Property Count; Asset Count",
        "formula": "",
    },
    {
        "metric_name": "Purchase Date",
        "metric_source": "extracted",
        "category": "Investment Basis",
        "definition": "Acquisition / closing date",
        "source_type": "General Information, Investment Summary",
        "data_nature": "mixed",
        "priority": "High",
        "core_question": "When was the deal closed?",
        "aliases": "Purchase Date; Acquisition Date; Closing Date; Close Date; Deal Close; Investment Date",
        "formula": "",
    },
    {
        "metric_name": "Exit Date",
        "metric_source": "extracted",
        "category": "Valuation & Returns",
        "definition": "Projected disposition / exit / sale date",
        "source_type": "General Information, Key UW Metrics, Returns",
        "data_nature": "projection",
        "priority": "High",
        "core_question": "When does the model exit?",
        "aliases": "Exit Date; Disposition Date; Sale Date; Reversion Date; Year of Exit; Disposition Year",
        "formula": "",
    },
    {
        "metric_name": "Interest Rate Spread",
        "metric_source": "extracted",
        "category": "Debt & Leverage",
        "definition": "Spread over the benchmark index (LIBOR / SOFR) for floating-rate debt",
        "source_type": "Debt, Debt Information",
        "data_nature": "projection",
        "priority": "High",
        "core_question": "What is the floating-rate spread?",
        "aliases": "Spread; Interest Rate Spread; Spread Over LIBOR; Spread Over SOFR; Spread Over Index; Loan Spread; Debt Spread",
        "formula": "",
    },
    {
        "metric_name": "Interest Rate Cap",
        "metric_source": "extracted",
        "category": "Debt & Leverage",
        "definition": "Strike rate of the interest rate cap hedge (max benchmark rate hedged)",
        "source_type": "Debt, Debt Information",
        "data_nature": "projection",
        "priority": "Medium",
        "core_question": "What's the cap on floating-rate exposure?",
        "aliases": "Interest Rate Cap; Rate Cap; Cap Strike; LIBOR Cap; SOFR Cap; Hedge Cap; Strike Rate",
        "formula": "",
    },
]


# ----------------------------------------------------------------------------
# ALIAS EXPANSIONS — updates to existing rows
# ----------------------------------------------------------------------------
ALIAS_UPDATES: dict[str, str] = {
    "Physical Occupancy": (
        "Physical Occupancy; Occupancy Rate; Physical Occ; Current Occupancy; "
        "Occupied Units; Physical Occ %; Occupancy %; Occ %; % Occupied; "
        "Occupancy at Acquisition; In-Place Occupancy; Stabilized Occupancy; Occupancy"
    ),
    "Interest-Only Period Remaining": (
        "Interest-Only Period Remaining; I/O Period; IO Period; Interest Only Period; "
        "I/O Months; IO Months; I/O Term; IO Term; Interest-Only Term; "
        "Interest Only; I/O; IO"
    ),
    "Interest Rate": (
        "Interest Rate; All-in Rate; All-in Interest Rate; Coupon; Coupon Rate; "
        "Pay Rate; Effective Rate; Loan Rate; Note Rate; Stated Rate; "
        "Fixed Rate; Floating Rate; Total Interest Rate"
    ),
    "Hold Period": (
        # Tightened: removed bare "Hold" and "Duration" which match unrelated cells.
        # Removed "Project Timeline" which matched NOI sheet timeline values.
        "Hold Period; Holding Period; Investment Period; Hold Term; "
        "Investment Horizon; Years Held; Hold (Years); Holding Period (Years); "
        "Total Hold; Total Hold Period"
    ),
    "CapEx Budget": (
        "Initial CapEx Budget; CapEx Budget; Total CapEx; Total Capital Expenditure; "
        "CapEx; Capital Expenditure; Capital Costs; Renovation Cost; Renovation Budget; "
        "Total Renovation Cost; Capital Plan; Total Capital Plan; CapEx Total; "
        "Capital Improvement Budget; Total Capital Improvements"
    ),
}


# ----------------------------------------------------------------------------
# PROPERTY-TYPE APPLICABILITY — by metric_name
# ----------------------------------------------------------------------------
# Empty/blank means "applies to all property types".
# Anything listed restricts the metric to those types.
APPLIES_TO_PROPERTY_TYPES: dict[str, str] = {
    # Total SF doesn't apply to Hotels (they track Keys, not building SF).
    # Listed for non-hotel asset types only.
    "Total SF": "Office; Multifamily; Industrial; Retail; Mixed-use; Self-Storage; Senior Living",
    # Parking applies to most types but not single-tenant industrial typically.
    # Leaving permissive for now.
}


def main():
    wb = load_workbook(CATALOG_PATH)
    ws = wb.active
    header_row = [c.value for c in ws[1]]
    name_col_idx = header_row.index("metric_name") + 1

    # Ensure new column exists
    if "applies_to_property_types" not in header_row:
        ws.cell(row=1, column=ws.max_column + 1, value="applies_to_property_types")
        header_row = [c.value for c in ws[1]]
    applies_col_idx = header_row.index("applies_to_property_types") + 1
    alias_col_idx   = header_row.index("aliases") + 1

    # 1) Apply alias updates to existing rows
    alias_updates_done = []
    for row in ws.iter_rows(min_row=2):
        name = row[name_col_idx - 1].value
        if name in ALIAS_UPDATES:
            row[alias_col_idx - 1].value = ALIAS_UPDATES[name]
            alias_updates_done.append(name)

    # 2) Apply property-type restrictions to existing rows
    pt_updates_done = []
    for row in ws.iter_rows(min_row=2):
        name = row[name_col_idx - 1].value
        if name in APPLIES_TO_PROPERTY_TYPES:
            row[applies_col_idx - 1].value = APPLIES_TO_PROPERTY_TYPES[name]
            pt_updates_done.append(name)

    # 3) Append new metric rows
    existing_names = {row[name_col_idx - 1].value for row in ws.iter_rows(min_row=2)}
    new_added = []
    for m in NEW_METRICS:
        if m["metric_name"] in existing_names:
            continue
        new_row = [None] * len(header_row)
        for k, v in m.items():
            if k in header_row:
                new_row[header_row.index(k)] = v
        ws.append(new_row)
        new_added.append(m["metric_name"])

    wb.save(CATALOG_PATH)
    print("Phase 1.5a updates applied:")
    print(f"  Alias updates       : {len(alias_updates_done)} ({', '.join(alias_updates_done)})")
    print(f"  Property-type rules : {len(pt_updates_done)} ({', '.join(pt_updates_done)})")
    print(f"  New metrics added   : {len(new_added)} ({', '.join(new_added)})")


if __name__ == "__main__":
    main()
