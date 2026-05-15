from pathlib import Path
import re
import json
import math
import datetime
import openpyxl
import pandas as pd

from flexible_extractor import scan_workbook_for_metric
from metric_catalog import load_metric_catalog


REPOSITORY_DIR = Path("repository")


# ---------------------------------------------------
# Shared helpers
# ---------------------------------------------------

def safe_float(value):
    """Convert a cell value to float, handling currency strings like '$1,234,567' and '-'."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            if math.isnan(value):
                return None
        except Exception:
            pass
        return float(value)
    try:
        cleaned = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
        if cleaned in ("", "-", "—", "n/a", "na"):
            return None
        return float(cleaned)
    except Exception:
        return None




def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def cell_addr(row, col):
    return openpyxl.utils.get_column_letter(col) + str(row)


def make_metric_id(name):
    text = name.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def find_label_cell(ws, search_text, prefer_last=True):
    """Returns (row, col) of the cell containing the label text, or (None, None)."""
    search_text = normalize_text(search_text)
    matches = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and search_text in normalize_text(cell.value):
                matches.append((cell.row, cell.column))
    if not matches:
        return None, None
    return matches[-1] if prefer_last else matches[0]


def find_total_column(ws, max_scan_rows=15):
    """
    Find the column that represents a total/annual/YTD figure.
    Searches header rows for keywords. Returns column index or None.
    """
    total_keywords = ["total", "annual", "ytd", "full year", "year total", "t12"]
    for row_num in range(1, max_scan_rows + 1):
        for col in range(1, ws.max_column + 1):
            cell_text = normalize_text(ws.cell(row=row_num, column=col).value)
            if any(kw in cell_text for kw in total_keywords):
                return col
    return None


def find_last_numeric_column(ws, row_num, start_col=2):
    """Find the rightmost column in a row that has a numeric value."""
    last_col = None
    for col in range(start_col, ws.max_column + 1):
        if safe_float(ws.cell(row=row_num, column=col).value) is not None:
            last_col = col
    return last_col


def sum_row_range(ws, row_num, start_col, end_col):
    total = 0
    count = 0
    for col in range(start_col, end_col + 1):
        val = safe_float(ws.cell(row=row_num, column=col).value)
        if val is not None:
            total += val
            count += 1
    return total if count > 0 else None


def make_metric(name, value, source_file, sheet, label_cell, value_cell,
                category="", definition="", method="playbook"):
    return {
        "metric_id": make_metric_id(name),
        "metric_name": name,
        "category": category,
        "definition": definition,
        "value": value,
        "source_file": Path(source_file).name,
        "sheet": sheet,
        "label_cell": label_cell,
        "value_cell": value_cell,
        "matched_alias": name,
        "confidence": "high",
        "match_method": method,
    }


def load_workbook(file_path):
    """Shared workbook loader — returns workbook or None on failure."""
    try:
        return openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return None


def get_sheets_to_scan(wb, relevant_tabs):
    """Return relevant_tabs intersected with actual sheet names, falling back to all sheets."""
    if relevant_tabs:
        matched = [s for s in relevant_tabs if s in wb.sheetnames]
        return matched if matched else wb.sheetnames
    return wb.sheetnames


# ---------------------------------------------------
# Financial statement playbook
# ---------------------------------------------------

REVENUE_LABELS = [
    "total operating revenue",
    "effective gross revenue",
    "total revenue",
    "gross revenue",
    "egi",
]

EXPENSE_LABELS = [
    "total operating expenses",
    "total expenses",
    "total opex",
    "operating expenses",
]

NOI_LABELS = [
    "net operating income",
    "noi",
]

EXPENSE_LINE_ITEMS = {
    "Repair & Maintenance Expense": ["total repair & maintenance", "repair & maintenance", "r&m"],
    "Utilities Expense":            ["total utilities", "utilities expense"],
    "Property Tax Expense":         ["total property tax", "property tax expense"],
    "Property Insurance Expense":   ["total property insurance", "property insurance"],
    "Property Management Expense":  ["total property management", "property management expense"],
    "Administrative Expense":       ["total administrative", "administrative expense"],
    "Leasing & Marketing Expense":  ["total lease & marketing", "leasing & marketing"],
    "Professional Service Expense": ["total professional service", "professional service"],
}


def _extract_row_value(ws, labels, total_col):
    """
    Given a list of label search strings, find the row and extract the value
    from total_col if available, otherwise sum all numeric columns in that row.
    Returns (value, label_cell_addr, value_cell_addr, matched_label).
    """
    for label in labels:
        row_num, label_col = find_label_cell(ws, label)
        if row_num is None:
            continue

        label_cell = cell_addr(row_num, label_col)
        data_start_col = label_col + 1

        # Try the dedicated total column first
        if total_col:
            val = safe_float(ws.cell(row=row_num, column=total_col).value)
            if val is not None:
                return val, label_cell, cell_addr(row_num, total_col), label
            # Total column header exists but this row is empty (e.g. file has
            # a "Total" header column that isn't populated) — fall through to sum

        # Sum all numeric columns after the label (excludes account codes / text)
        last_col = find_last_numeric_column(ws, row_num, start_col=data_start_col)
        if last_col:
            # Exclude the total column itself from the sum range to avoid double-counting
            sum_end = (total_col - 1) if (total_col and total_col <= last_col) else last_col
            val = sum_row_range(ws, row_num, data_start_col, sum_end)
            if val is not None:
                return (
                    val, label_cell,
                    f"{cell_addr(row_num, data_start_col)}:{cell_addr(row_num, sum_end)}",
                    label
                )

    return None, None, None, None


def extract_financial_statement(file_path, relevant_tabs=None):
    """
    Extracts NOI, revenue, expenses from a financial statement workbook.
    Uses the total/annual column if found, otherwise sums monthly columns.
    Also captures monthly series for trend analysis.
    """
    metrics = []
    wb = load_workbook(file_path)
    if wb is None:
        return metrics

    for sheet_name in get_sheets_to_scan(wb, relevant_tabs):
        ws = wb[sheet_name]
        total_col = find_total_column(ws)

        rev_val, rev_lc, rev_vc, _ = _extract_row_value(ws, REVENUE_LABELS, total_col)
        exp_val, exp_lc, exp_vc, _ = _extract_row_value(ws, EXPENSE_LABELS, total_col)
        noi_val, noi_lc, noi_vc, _ = _extract_row_value(ws, NOI_LABELS, total_col)

        if rev_val is not None:
            metrics.append(make_metric(
                "Revenue", rev_val, file_path, sheet_name, rev_lc, rev_vc,
                category="Operating Performance", definition="Total operating revenue"
            ))

        if exp_val is not None:
            metrics.append(make_metric(
                "Operating Expenses", exp_val, file_path, sheet_name, exp_lc, exp_vc,
                category="Operating Performance", definition="Total operating expenses"
            ))

        if noi_val is not None:
            metrics.append(make_metric(
                "NOI", noi_val, file_path, sheet_name, noi_lc, noi_vc,
                category="Operating Performance", definition="Net Operating Income"
            ))
        elif rev_val is not None and exp_val is not None:
            metrics.append(make_metric(
                "NOI", rev_val - abs(exp_val), file_path, sheet_name, "", "",
                category="Operating Performance",
                definition="Net Operating Income (calculated: Revenue - Expenses)",
                method="calculated"
            ))

        for metric_name, labels in EXPENSE_LINE_ITEMS.items():
            val, lc, vc, _ = _extract_row_value(ws, labels, total_col)
            if val is not None:
                metrics.append(make_metric(
                    metric_name, val, file_path, sheet_name, lc, vc,
                    category="Expense Detail"
                ))

        # --- Monthly series (for trend analysis) ---
        # Collect all numeric columns in the revenue row, excluding the total column.
        # The resulting arrays power core question 6: "Is risk increasing or decreasing?"
        for series_name, labels in [("Monthly Revenue", REVENUE_LABELS),
                                     ("Monthly Expenses", EXPENSE_LABELS),
                                     ("Monthly NOI", NOI_LABELS)]:
            for label in labels:
                row_num, label_col = find_label_cell(ws, label)
                if row_num is None:
                    continue

                monthly = []
                for col in range(label_col + 1, ws.max_column + 1):
                    if total_col and col == total_col:
                        continue
                    val = safe_float(ws.cell(row=row_num, column=col).value)
                    if val is not None:
                        monthly.append(val)

                if len(monthly) >= 6:
                    metrics.append(make_metric(
                        series_name, monthly, file_path, sheet_name,
                        cell_addr(row_num, label_col), "",
                        category="Operating Performance",
                        definition=f"Monthly {series_name.lower().replace('monthly ', '')} series",
                        method="monthly_series"
                    ))
                break

        if metrics:
            break

    return metrics


# ---------------------------------------------------
# Debt metric calculations
# ---------------------------------------------------

def _get_val(metrics, *names):
    """Retrieve the value of the first matching metric name from a list."""
    name_set = {n.lower() for n in names}
    for m in metrics:
        if m["metric_name"].lower() in name_set:
            v = m["value"]
            return v if isinstance(v, (int, float)) else None
    return None


def calculate_derived_debt_metrics(metrics, source_file, sheet):
    """
    Derives DSCR, Debt Yield, and Debt Service Constant from already-extracted
    values when they were not found directly as labeled cells.
    """
    derived = []
    extracted_names = {m["metric_name"].lower() for m in metrics}

    noi = _get_val(metrics, "NOI", "Original NOI", "Forward NOI", "Annual NOI")
    debt_service = _get_val(metrics, "Debt Service")
    loan_amount = _get_val(metrics, "Debt Amount", "Initial Debt", "Total Debt", "Loan Balance")
    loan_balance = _get_val(metrics, "Loan Balance", "Debt Amount", "Initial Debt", "Total Debt")

    # DSCR = NOI / Annual Debt Service
    if "dscr / debt coverage ratio" not in extracted_names:
        if noi and debt_service and debt_service != 0:
            derived.append(make_metric(
                "DSCR / Debt Coverage Ratio",
                round(noi / abs(debt_service), 3),
                source_file, sheet, "", "",
                category="Debt & Leverage",
                definition="NOI / Annual Debt Service (calculated)",
                method="calculated"
            ))

    # Debt Yield = NOI / Loan Balance
    if "debt yield" not in extracted_names:
        if noi and loan_balance and loan_balance != 0:
            derived.append(make_metric(
                "Debt Yield",
                round(noi / abs(loan_balance), 4),
                source_file, sheet, "", "",
                category="Debt & Leverage",
                definition="NOI / Loan Balance (calculated)",
                method="calculated"
            ))

    # Debt Service Constant = Annual Debt Service / Loan Amount
    if "debt service constant" not in extracted_names:
        if debt_service and loan_amount and loan_amount != 0:
            derived.append(make_metric(
                "Debt Service Constant",
                round(abs(debt_service) / abs(loan_amount), 4),
                source_file, sheet, "", "",
                category="Debt & Leverage",
                definition="Annual Debt Service / Loan Amount (calculated)",
                method="calculated"
            ))

    # Current LTV = Loan Balance / Value (if value is available)
    if "current ltv" not in extracted_names:
        value = _get_val(metrics, "Exit Value", "Value", "Implied Value")
        if loan_balance and value and value != 0:
            derived.append(make_metric(
                "Current LTV",
                round(abs(loan_balance) / abs(value), 4),
                source_file, sheet, "", "",
                category="Debt & Leverage",
                definition="Loan Balance / Current Value (calculated)",
                method="calculated"
            ))

    # Going-in Cap Rate = Original NOI / Purchase Price
    if "going-in cap rate" not in extracted_names:
        orig_noi = _get_val(metrics, "Original NOI")
        purchase_price = _get_val(metrics, "Purchase Price")
        if orig_noi and purchase_price and purchase_price != 0:
            derived.append(make_metric(
                "Going-in Cap Rate",
                round(orig_noi / purchase_price, 4),
                source_file, sheet, "", "",
                category="Valuation & Returns",
                definition="Original NOI / Purchase Price (calculated)",
                method="calculated"
            ))

    return derived


# ---------------------------------------------------
# Proforma capital-spend scanner (shared by UW and BP playbooks)
# ---------------------------------------------------

def extract_capex_spent_from_proforma(wb, file_path):
    """
    Scan a workbook's Proforma sheet for actual TI/LC and CapEx spend rows,
    sum the values that sit under columns marked 'Actual', and return CapEx Spent
    and a forecast counterpart. Returns [] if no Proforma sheet or no spend rows.

    Looks for a header row (rows 1-8) where cells contain 'Actual' or 'Forecast'
    period markers — this is how the BP/UW Proforma indicates which periods are
    realized vs. projected.
    """
    if "Proforma" not in wb.sheetnames:
        return []

    ws = wb["Proforma"]
    metrics = []

    # Find the row marking columns as Actual vs Forecast (typically rows 2-6)
    actual_cols, forecast_cols = [], []
    for r in range(1, min(9, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            marker = normalize_text(ws.cell(row=r, column=c).value)
            if marker == "actual":
                actual_cols.append(c)
            elif marker == "forecast":
                forecast_cols.append(c)
        if actual_cols or forecast_cols:
            break

    if not actual_cols and not forecast_cols:
        return []

    # Spend-row labels to scan (TI, LC, and explicit CapEx lines)
    spend_labels = [
        ("ti", ["tenant improvement", "ti spend", "ti / lc", "ti/lc"]),
        ("lc", ["leasing commission", "lc spend"]),
        ("capex", ["capex", "capital expenditure", "capital spend"]),
    ]

    actual_total = 0.0
    forecast_total = 0.0
    found_any = False
    label_cells = []

    for _, labels in spend_labels:
        for label in labels:
            row_num, label_col = find_label_cell(ws, label, prefer_last=False)
            if row_num is None:
                continue

            label_cells.append(cell_addr(row_num, label_col))

            for c in actual_cols:
                v = safe_float(ws.cell(row=row_num, column=c).value)
                if v is not None:
                    actual_total += abs(v)
                    found_any = True
            for c in forecast_cols:
                v = safe_float(ws.cell(row=row_num, column=c).value)
                if v is not None:
                    forecast_total += abs(v)
                    found_any = True
            break  # only take the first matching label per category

    if not found_any:
        return []

    label_ref = label_cells[0] if label_cells else ""

    if actual_total > 0:
        metrics.append(make_metric(
            "CapEx / TI/LC Spent", round(actual_total, 2),
            file_path, "Proforma", label_ref, "",
            category="CapEx & Capital Allocation",
            definition="Sum of TI + LC + CapEx in Proforma 'Actual' columns",
            method="proforma_actual_sum"
        ))

    if forecast_total > 0:
        metrics.append(make_metric(
            "CapEx / TI/LC Forecast Remaining", round(forecast_total, 2),
            file_path, "Proforma", label_ref, "",
            category="CapEx & Capital Allocation",
            definition="Sum of TI + LC + CapEx in Proforma 'Forecast' columns",
            method="proforma_forecast_sum"
        ))

    return metrics


# ---------------------------------------------------
# Underwriting playbook
# ---------------------------------------------------

UNDERWRITING_TARGETS = [
    # --- Basis & valuation ---
    ("Purchase Price",    ["purchase price", "acquisition price"],                          "Valuation & Returns", "Acquisition purchase price"),
    ("Total Basis",       ["total going in basis", "total basis", "cost basis",
                           "all-in basis", "going in basis"],                               "Valuation & Returns", "All-in cost basis"),
    ("Going-in Cap Rate", ["going-in cap", "going in cap", "in-place cap",
                           "acquisition cap rate"],                                          "Valuation & Returns", "Cap rate at acquisition"),
    ("Exit Cap Rate",     ["take out cap rate", "exit cap", "reversion cap",
                           "terminal cap", "cap rate"],                                      "Valuation & Returns", "Projected cap rate at exit"),
    ("Exit Value",        ["exit value", "reversion value", "sale price"],                  "Valuation & Returns", "Projected sale price at exit"),
    ("Yield on Cost",     ["yield on cost", "yoc", "stabilized yield"],                     "Valuation & Returns", "Stabilized NOI / Total Basis"),

    # --- Operating ---
    ("Original NOI",      ["original noi", "in-place noi", "t12 noi", "going-in noi"],     "Operating Performance", "NOI at acquisition"),
    ("Forward NOI",       ["forward noi", "stabilized noi", "projected noi"],               "Operating Performance", "Projected stabilized NOI"),

    # --- Debt terms ---
    ("Debt Amount",       ["initial funding", "debt amount", "loan amount", "initial debt",
                           "senior loan", "first mortgage", "total loan"],                  "Debt & Leverage", "Total loan amount at closing"),
    ("Total Debt",        ["total debt"],                                                   "Debt & Leverage", "Total debt including future fundings"),
    ("Loan Balance",      ["loan balance", "outstanding balance", "current balance",
                           "remaining balance", "outstanding debt", "debt balance"],        "Debt & Leverage", "Current outstanding loan balance"),
    ("Original LTV",      ["original ltv", "initial ltv", "ltv at closing",
                           "ltv", "loan to value", "loan-to-value"],                        "Debt & Leverage", "LTV at acquisition"),
    ("Interest Rate",     ["interest rate spread", "interest rate", "note rate",
                           "coupon rate", "spread"],                                        "Debt & Leverage", "Loan interest rate / spread"),
    ("Loan Maturity",     ["term", "loan maturity", "maturity date", "loan term"],          "Debt & Leverage", "Loan maturity date or term"),
    ("LTC",               ["ltc", "loan to cost", "loan-to-cost"],                         "Debt & Leverage", "Loan to cost ratio"),

    # --- Coverage & leverage ratios ---
    ("DSCR / Debt Coverage Ratio", ["dscr", "dscr check", "debt service coverage",
                                     "debt coverage", "dsc ratio"],                        "Debt & Leverage", "NOI / Annual Debt Service"),
    ("Debt Yield",        ["debt yield", "dy"],                                             "Debt & Leverage", "NOI / Loan Balance"),
    ("Debt Service Constant", ["debt service constant", "mortgage constant",
                                "loan constant", "dsc constant"],                           "Debt & Leverage", "Annual Debt Service / Loan Amount"),
    ("Break-even Occupancy (Monthly)", ["break-even occupancy", "breakeven occupancy",
                                         "break even occ", "breakeven occ"],               "Debt & Leverage", "Occupancy needed to cover all fixed costs"),

    # --- CapEx / TI/LC (treated as one capital bucket) ---
    # Note: "ti/lc" and "ti / lc" both searched to handle spacing variants in models
    ("CapEx / TI/LC Budget", ["ti/lc", "ti / lc", "ti & lc", "ti and lc",
                               "capex / ti / lc", "capex/ti/lc",
                               "capex", "capital expenditure", "capital costs",
                               "initial capex", "capex budget",
                               "tenant improvement", "leasing commission"],                "CapEx & Capital Allocation", "Total capital budget including CapEx, TI, and LC"),
    ("CapEx / TI/LC Spent",  ["capex spent", "capital spent", "capex to date",
                               "ti spent", "lc spent", "ti/lc spent"],                  "CapEx & Capital Allocation", "Capital deployed to date including TI/LC"),
    ("CapEx Remaining",      ["capex remaining", "remaining capex",
                               "remaining capital", "cost to complete"],                "CapEx & Capital Allocation", "Remaining capital budget"),
    ("ROI on CapEx / TI/LC", ["roi on capex", "capex roi", "return on capex",
                               "return on capital", "ti roi", "lc roi"],                "CapEx & Capital Allocation", "Return on deployed capital including TI/LC"),
    ("Yield on Incremental Cost", ["yield on incremental", "incremental yield",
                                    "incremental noi", "stabilized yield on cost"],     "CapEx & Capital Allocation", "Incremental NOI / CapEx + TI/LC spend"),

    # --- Returns ---
    ("Unlevered IRR",     ["unlevered irr", "property irr"],                            "Valuation & Returns", "Unlevered internal rate of return"),
    ("Levered IRR",       ["levered irr", "equity irr"],                                "Valuation & Returns", "Levered internal rate of return"),
    ("Equity Multiple",   ["equity multiple", "moic"],                                  "Valuation & Returns", "Equity return multiple"),
]


def extract_underwriting_summary(file_path, relevant_tabs=None):
    """
    Extracts acquisition basis, debt, IRR, and exit assumptions from an underwriting model.
    """
    metrics = []
    wb = load_workbook(file_path)
    if wb is None:
        return metrics

    seen_names = set()

    for sheet_name in get_sheets_to_scan(wb, relevant_tabs):
        ws = wb[sheet_name]

        for metric_name, labels, category, definition in UNDERWRITING_TARGETS:
            if metric_name in seen_names:
                continue

            for label in labels:
                row_num, label_col = find_label_cell(ws, label)
                if row_num is None:
                    continue

                for offset in range(1, 8):
                    val = safe_float(ws.cell(row=row_num, column=label_col + offset).value)
                    if val is not None and val != 0:
                        metrics.append(make_metric(
                            metric_name, val, file_path, sheet_name,
                            cell_addr(row_num, label_col),
                            cell_addr(row_num, label_col + offset),
                            category=category, definition=definition
                        ))
                        seen_names.add(metric_name)
                        break
                break

    # Derive coverage ratios not found directly as labeled cells
    source_sheet = metrics[0]["sheet"] if metrics else ""
    metrics += calculate_derived_debt_metrics(metrics, file_path, source_sheet)

    # Sum TI/LC/CapEx spend rows from the Proforma actual columns
    metrics += extract_capex_spent_from_proforma(wb, file_path)

    return metrics


# ---------------------------------------------------
# Business plan playbook
# ---------------------------------------------------

BP_RETURN_TARGETS = [
    # --- Returns ---
    ("BP Unlevered IRR",   ["unlevered irr", "property irr"],       "Valuation & Returns", "Business plan unlevered IRR"),
    ("BP Levered IRR",     ["levered irr", "equity irr"],           "Valuation & Returns", "Business plan levered IRR"),
    ("BP Equity Multiple", ["equity multiple", "moic"],             "Valuation & Returns", "Business plan equity multiple"),
    ("BP Exit Cap Rate",   ["exit cap", "reversion cap"],           "Valuation & Returns", "Business plan exit cap rate"),
    ("BP Exit Value",      ["exit value", "reversion value"],       "Valuation & Returns", "Business plan exit value"),

    # --- Debt (in-place and refinance) ---
    ("Debt Service",       ["annual debt service", "debt service",
                            "principal and interest", "p&i"],       "Debt & Leverage", "Annual debt service payment"),
    ("Refinance DSCR",     ["refi dscr", "refinance dscr",
                            "exit dscr", "reversion dscr"],         "Debt & Leverage", "Projected DSCR at refinance or exit"),
    ("DSCR / Debt Coverage Ratio", ["dscr", "debt service coverage",
                                     "debt coverage"],              "Debt & Leverage", "NOI / Annual Debt Service"),
    ("Debt Yield",         ["debt yield", "dy"],                    "Debt & Leverage", "NOI / Loan Balance"),
    ("Debt Service Constant", ["debt service constant",
                                "mortgage constant", "loan constant"], "Debt & Leverage", "Annual Debt Service / Loan Amount"),
    ("Loan Balance",       ["loan balance", "outstanding balance",
                            "remaining balance", "outstanding debt"], "Debt & Leverage", "Outstanding loan balance"),
    ("Break-even Occupancy (Monthly)", ["break-even occupancy",
                                         "breakeven occupancy",
                                         "break even occ"],         "Debt & Leverage", "Occupancy needed to cover fixed costs"),
]


def extract_business_plan_summary(file_path, relevant_tabs=None):
    """
    Extracts projected operating metrics and return expectations from a business plan.
    Reuses the financial statement playbook for operating lines, adds return metrics.
    """
    metrics = extract_financial_statement(file_path, relevant_tabs)

    wb = load_workbook(file_path)
    if wb is None:
        return metrics

    seen_names = set()

    for sheet_name in get_sheets_to_scan(wb, relevant_tabs):
        ws = wb[sheet_name]

        for metric_name, labels, category, definition in BP_RETURN_TARGETS:
            if metric_name in seen_names:
                continue

            for label in labels:
                row_num, label_col = find_label_cell(ws, label)
                if row_num is None:
                    continue

                for offset in range(1, 8):
                    val = safe_float(ws.cell(row=row_num, column=label_col + offset).value)
                    if val is not None and val != 0:
                        metrics.append(make_metric(
                            metric_name, val, file_path, sheet_name,
                            cell_addr(row_num, label_col),
                            cell_addr(row_num, label_col + offset),
                            category=category, definition=definition
                        ))
                        seen_names.add(metric_name)
                        break
                break

    # Derive coverage ratios not found directly
    source_sheet = metrics[0]["sheet"] if metrics else ""
    metrics += calculate_derived_debt_metrics(metrics, file_path, source_sheet)

    # Sum TI/LC/CapEx spend rows from the Proforma actual columns
    metrics += extract_capex_spent_from_proforma(wb, file_path)

    return metrics


# ---------------------------------------------------
# Rent roll playbook
# ---------------------------------------------------

def _find_header_row(ws, keywords, max_scan=25):
    """Find the row where at least 2 of the given keywords appear across columns."""
    for row_num in range(1, max_scan + 1):
        row_text = " ".join(
            normalize_text(ws.cell(row=row_num, column=col).value)
            for col in range(1, min(ws.max_column + 1, 25))
        )
        if sum(1 for kw in keywords if kw in row_text) >= 2:
            return row_num
    return None


def _find_col_by_header(ws, header_row, keywords):
    """Find the first column whose header matches any keyword."""
    for col in range(1, ws.max_column + 1):
        cell_text = normalize_text(ws.cell(row=header_row, column=col).value)
        if any(kw in cell_text for kw in keywords):
            return col
    return None


def extract_rent_roll(file_path, relevant_tabs=None):
    """
    Extracts occupancy, tenant count, and tenant concentration from a rent roll.
    Detects the header row, reads tenant rows, and derives summary metrics.
    """
    metrics = []
    wb = load_workbook(file_path)
    if wb is None:
        return metrics

    for sheet_name in get_sheets_to_scan(wb, relevant_tabs):
        ws = wb[sheet_name]

        header_row = _find_header_row(ws, ["tenant", "sf", "rent", "expir"])
        if header_row is None:
            continue

        tenant_col = _find_col_by_header(ws, header_row, ["tenant", "name"])
        unit_col   = _find_col_by_header(ws, header_row, ["unit", "suite", "space", "#"])
        sf_col     = _find_col_by_header(ws, header_row, ["sf", "sqft", "sq ft", "rsf", "nrsf", "area", "size"])
        # Prefer annual/monthly total over per-sf rent; avoid "Rent Reset" column
        rent_col   = (
            _find_col_by_header(ws, header_row, ["annual total", "monthly total", "annual rent",
                                                  "base rent", "psf annual", "psf monthly"])
            or _find_col_by_header(ws, header_row, ["rent/sf", "rent psf"])
        )
        expiry_col = _find_col_by_header(ws, header_row, ["lease end", "expir", "expiry", "end date"])

        total_sf = 0
        occupied_sf = 0
        tenants = []

        for row_num in range(header_row + 1, ws.max_row + 1):
            tenant_val = ws.cell(row=row_num, column=tenant_col).value if tenant_col else None
            sf_val     = safe_float(ws.cell(row=row_num, column=sf_col).value) if sf_col else None

            if tenant_val is None and sf_val is None:
                continue

            tenant_text = normalize_text(tenant_val)

            # Skip summary/total rows
            if any(x in tenant_text for x in ["total", "subtotal"]):
                continue

            # Skip rows with no unit number and no lease date — these are summary
            # or footnote rows, not individual tenant records
            unit_val = ws.cell(row=row_num, column=unit_col).value if unit_col else None
            expiry_val_check = ws.cell(row=row_num, column=expiry_col).value if expiry_col else None
            if unit_val is None and not isinstance(expiry_val_check, datetime.datetime):
                continue

            if sf_val is None:
                continue

            total_sf += sf_val

            rent_val   = safe_float(ws.cell(row=row_num, column=rent_col).value) if rent_col else None
            expiry_val = ws.cell(row=row_num, column=expiry_col).value if expiry_col else None

            # Use lease-end datetime as the occupancy signal — more reliable than
            # name presence, since many models leave tenant names blank for occupied units
            has_lease  = isinstance(expiry_val, datetime.datetime)
            is_vacant  = "vacant" in tenant_text or not has_lease

            if not is_vacant:
                occupied_sf += sf_val
                tenants.append({
                    "tenant": tenant_val,
                    "sf": sf_val,
                    "rent": rent_val,
                    "expiry": expiry_val,
                })

        if not tenants:
            continue

        if total_sf > 0:
            metrics.append(make_metric(
                "Occupancy", round(occupied_sf / total_sf, 4),
                file_path, sheet_name, "", "",
                category="Leasing", definition="Occupied SF / Total SF"
            ))
            metrics.append(make_metric(
                "Occupied SF", occupied_sf, file_path, sheet_name, "", "",
                category="Leasing", definition="Total occupied square footage"
            ))
            metrics.append(make_metric(
                "Total SF", total_sf, file_path, sheet_name, "", "",
                category="Leasing", definition="Total rentable square footage"
            ))

        metrics.append(make_metric(
            "Tenant Count", len(tenants), file_path, sheet_name, "", "",
            category="Leasing", definition="Number of occupied tenants on rent roll"
        ))

        if total_sf > 0 and tenants:
            top = max(tenants, key=lambda x: x["sf"] or 0)
            top_sf = top["sf"] or 0
            metrics.append(make_metric(
                "Tenant Concentration", round(top_sf / total_sf, 4),
                file_path, sheet_name, "", "",
                category="Leasing",
                definition=f"Largest tenant ({top['tenant']}) as % of total SF"
            ))

        # WALT — SF-weighted average remaining lease term in years
        today = datetime.datetime.now()
        weighted_months = 0.0
        weighted_sf = 0.0
        for t in tenants:
            if isinstance(t["expiry"], datetime.datetime) and t["sf"]:
                remaining_days = (t["expiry"] - today).days
                if remaining_days > 0:
                    weighted_months += (remaining_days / 30.44) * t["sf"]
                    weighted_sf += t["sf"]

        if weighted_sf > 0:
            walt_years = (weighted_months / weighted_sf) / 12
            metrics.append(make_metric(
                "WALT", round(walt_years, 2),
                file_path, sheet_name, "", "",
                category="Leasing",
                definition="SF-weighted average remaining lease term, in years"
            ))

        # Near-term rollover — % of occupied SF expiring within 24 months
        rollover_sf = 0.0
        for t in tenants:
            if isinstance(t["expiry"], datetime.datetime) and t["sf"]:
                months_remaining = (t["expiry"] - today).days / 30.44
                if 0 < months_remaining <= 24:
                    rollover_sf += t["sf"]

        if occupied_sf > 0:
            metrics.append(make_metric(
                "Rollover (24mo)", round(rollover_sf / occupied_sf, 4),
                file_path, sheet_name, "", "",
                category="Leasing",
                definition="% of occupied SF rolling within 24 months"
            ))

        break

    return metrics


# ---------------------------------------------------
# Fallback: flexible alias scan
# ---------------------------------------------------

def run_flexible_scan(file_path, relevant_tabs=None):
    """Fallback for unclassified or unsupported file types — uses alias-based scanning."""
    catalog = load_metric_catalog()
    results = []
    for metric in catalog:
        match = scan_workbook_for_metric(file_path, metric, relevant_tabs=relevant_tabs)
        if match:
            results.append(match)
    return results


# ---------------------------------------------------
# Router
# ---------------------------------------------------

def extract_document(file_path, document_type, relevant_tabs=None):
    """Route a single file to the appropriate extraction playbook."""
    dt = normalize_text(document_type)

    if dt == "financial_statement_actuals":
        return extract_financial_statement(file_path, relevant_tabs)
    elif dt == "acquisition_underwriting":
        return extract_underwriting_summary(file_path, relevant_tabs)
    elif dt == "business_plan":
        return extract_business_plan_summary(file_path, relevant_tabs)
    elif dt == "rent_roll":
        return extract_rent_roll(file_path, relevant_tabs)
    else:
        return run_flexible_scan(file_path, relevant_tabs)


# ---------------------------------------------------
# Main entry point
# ---------------------------------------------------

def extract_all_documents(upload_dir="uploads", classification_result=None):
    """
    Run playbook extraction across all uploaded files using classification context.
    Returns a result dict in the same schema as scan_uploaded_files().
    """
    upload_dir = Path(upload_dir)
    REPOSITORY_DIR.mkdir(exist_ok=True)

    file_map = {}
    if classification_result:
        for item in classification_result.get("classifications", []):
            fn = item.get("file_name")
            if fn:
                file_map[fn] = item

    excel_files = list(upload_dir.glob("*.xlsx")) + list(upload_dir.glob("*.xlsm"))
    all_extracted = []

    for file_path in excel_files:
        item = file_map.get(file_path.name, {})
        document_type  = item.get("document_type", "unknown")
        relevant_tabs  = item.get("relevant_tabs") or None
        extracted = extract_document(file_path, document_type, relevant_tabs)
        all_extracted.extend(extracted)

    # Deduplicate by metric_id, keeping first occurrence (playbook results win)
    seen = set()
    deduped = []
    for m in all_extracted:
        if m["metric_id"] not in seen:
            seen.add(m["metric_id"])
            deduped.append(m)

    catalog = load_metric_catalog()
    extracted_ids = {m["metric_id"] for m in deduped}
    missing = [
        {
            "metric_id": m["metric_id"],
            "metric_name": m["metric_name"],
            "category": m["category"],
            "definition": m["definition"],
            "source": m.get("source", ""),
            "priority": m.get("priority", "medium"),
            "aliases": m.get("aliases", []),
            "status": "missing",
        }
        for m in catalog
        if m["metric_id"] not in extracted_ids
    ]

    result = {
        "status": "success",
        "total_metrics": len(catalog),
        "extracted_count": len(deduped),
        "missing_count": len(missing),
        "extracted_metrics": deduped,
        "missing_metrics": missing,
    }

    with open(REPOSITORY_DIR / "document_extraction_result.json", "w") as f:
        json.dump(result, f, indent=2, default=str)

    pd.DataFrame(deduped).to_csv(REPOSITORY_DIR / "extracted_metrics_report.csv", index=False)
    pd.DataFrame(missing).to_csv(REPOSITORY_DIR / "missing_metrics_report.csv", index=False)

    return result


if __name__ == "__main__":
    result = extract_all_documents()
    print(f"Total metrics:  {result['total_metrics']}")
    print(f"Extracted:      {result['extracted_count']}")
    print(f"Missing:        {result['missing_count']}")
