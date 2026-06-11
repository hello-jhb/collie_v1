"""
financial_model_parser.py — table-centric reading of real estate models.

The extraction engine is metric-centric ("find NOI, find CapEx, find IRR"), but
institutional models are TABLE-centric. A human reads a cash-flow table like
this:

    "This is a MONTHLY roll-up. The date header across the top sets the timeline.
     Every line item below it inherits that timeline."

This module makes the engine read the same way. It finds tables on each sheet,
establishes each table's period axis + periodicity from its DATE HEADER ROW
(read top → right), then extracts the line items beneath (read top → bottom) —
each row inheriting the table's periodicity. Downstream consumers (periodicity
detection, AAM, formula trace, snapshot, deep dives) read structured tables
instead of guessing per metric.

Output per table:
    {
      "sheet", "title", "table_type",      # e.g. "cashflow_rollup"
      "periodicity",                        # monthly | quarterly | annual | ...
      "period_axis": "horizontal",
      "header_row", "label_col",
      "date_headers": [iso, ...],
      "period_cols": [col, ...],
      "rows": [ {"label", "row", "values", "values_by_period"} ],
    }

Deterministic (openpyxl only) — no GPT, no API key.
"""
from __future__ import annotations

import datetime as _dt
import logging
import statistics
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

log = logging.getLogger("fb.model_parser")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.parser] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

MODEL_PARSER_VERSION = "2026-06-10.1"

# Bounds — schedules live near the top; skip giant raw-data source tabs.
_MAX_SCAN_ROWS = 160
_MAX_SCAN_COLS = 140
_MAX_SHEET_ROWS = 600          # skip sheets bigger than this (raw P&L dumps)
_MIN_PERIODS = 4               # a period axis needs at least this many columns
_MAX_BODY_GAP = 2             # consecutive blank rows that end a table body


# ---------------------------------------------------------------------------
# Period-axis detection (read top → right)
# ---------------------------------------------------------------------------

def _periodicity_from_days(median_days: float) -> str | None:
    if median_days <= 0:
        return None
    if median_days <= 10:
        return "weekly"
    if median_days <= 45:
        return "monthly"
    if median_days <= 135:
        return "quarterly"
    if median_days <= 250:
        return "semiannual"
    if median_days <= 400:
        return "annual"
    return None


def _clean_axis(values: list[_dt.date]) -> str | None:
    """
    A real period axis is strictly increasing with consistent spacing. Returns
    its periodicity, or None if the sequence isn't a clean timeline (filters out
    rate-reset rows, scattered debt dates, duplicated headers, etc.).
    """
    if len(values) < _MIN_PERIODS:
        return None
    if not all(values[i] < values[i + 1] for i in range(len(values) - 1)):
        return None
    deltas = [(values[i + 1] - values[i]).days for i in range(len(values) - 1)]
    med = statistics.median(deltas)
    if med <= 0:
        return None
    consistent = sum(1 for d in deltas if 0.5 * med <= d <= 1.6 * med)
    if consistent < 0.7 * len(deltas):
        return None
    return _periodicity_from_days(med)


def _as_date(v: Any) -> _dt.date | None:
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    return None


def _gcell(grid: list[tuple], r: int, c: int) -> Any:
    """1-indexed lookup into an in-memory sheet grid (list of value tuples)."""
    if 1 <= r <= len(grid):
        row = grid[r - 1]
        if 1 <= c <= len(row):
            return row[c - 1]
    return None


def _row_width(grid: list[tuple], r: int) -> int:
    return len(grid[r - 1]) if 1 <= r <= len(grid) else 0


def _detect_header(grid: list[tuple], row: int) -> dict | None:
    """
    If `row` is a period-axis header, return its descriptor:
      {periodicity, period_cols, date_headers}.  Else None.

    Handles two header shapes:
      - datetime headers (monthly/quarterly/annual schedules), and
      - bare year-integer headers (2018, 2019, 2020 — annual roll-ups).
    """
    width = _row_width(grid, row)
    cells = grid[row - 1] if 1 <= row <= len(grid) else ()

    # 1) datetime axis
    dated = [(c + 1, _as_date(v)) for c, v in enumerate(cells)]
    dated = [(c, d) for c, d in dated if d]
    if len(dated) >= _MIN_PERIODS:
        cols = [c for c, _ in dated]
        dates = [d for _, d in dated]
        period = _clean_axis(dates)
        if period:
            return {
                "periodicity": period,
                "period_cols": cols,
                "date_headers": [d.isoformat() for d in dates],
            }

    # 2) bare year-integer axis (e.g. 2018, 2019, 2020, 2021)
    years = [(c + 1, v) for c, v in enumerate(cells)
             if isinstance(v, int) and 1990 <= v <= 2100]
    if len(years) >= _MIN_PERIODS:
        cols = [c for c, _ in years]
        vals = [y for _, y in years]
        if all(vals[i] < vals[i + 1] for i in range(len(vals) - 1)) and all(
            vals[i + 1] - vals[i] == 1 for i in range(len(vals) - 1)
        ):
            return {
                "periodicity": "annual",
                "period_cols": cols,
                "date_headers": [f"{y}-12-31" for y in vals],
            }
    return None


# ---------------------------------------------------------------------------
# Body extraction (read top → bottom) + classification
# ---------------------------------------------------------------------------

def _label_col(grid: list[tuple], header_row: int, first_period_col: int) -> int:
    """Find the column holding row labels — leftmost text column before the axis."""
    last = min(header_row + 12, len(grid))
    for c in range(1, first_period_col):
        for r in range(header_row + 1, last + 1):
            v = _gcell(grid, r, c)
            if isinstance(v, str) and v.strip():
                return c
    return 1


def _extract_rows(grid: list[tuple], header_row: int, label_col: int,
                  period_cols: list[int], date_headers: list[str]) -> list[dict]:
    rows: list[dict] = []
    blanks = 0
    r = header_row + 1
    last = min(len(grid), _MAX_SCAN_ROWS)
    while r <= last:
        label = _gcell(grid, r, label_col)
        label = label.strip() if isinstance(label, str) else None
        values = [_gcell(grid, r, c) for c in period_cols]
        has_num = any(isinstance(v, (int, float)) for v in values)

        # Stop if we hit another period-axis header (a new table starts here).
        if _detect_header(grid, r):
            break

        if not label and not has_num:
            blanks += 1
            if blanks >= _MAX_BODY_GAP:
                break
            r += 1
            continue
        blanks = 0

        if label and has_num:
            rows.append({
                "label": label,
                "row": r,
                "values": values,
                "values_by_period": {
                    date_headers[i]: values[i] for i in range(len(values))
                },
            })
        r += 1
    return rows


_TYPE_RULES = [
    ("rent_roll",        ("rent roll",)),
    ("sources_uses",     ("sources & uses", "sources and uses", "sources & u")),
    ("return_waterfall", ("waterfall", "promote", "distribution")),
    ("debt_schedule",    ("debt schedule", "amortization", "loan schedule")),
    ("capex_schedule",   ("capex schedule", "capital schedule", "capex plan")),
    ("lease_schedule",   ("lease expiration", "lease schedule", "rollover")),
    ("noi_schedule",     ("noi schedule", "noi monthly", "ttm noi")),
    ("cashflow_rollup",  ("roll-up", "roll up", "rollup", "cash flow", "cashflow",
                          "proforma", "pro forma")),
]


def _classify(title: str | None, labels: list[str], periodicity: str) -> str:
    blob = (title or "").lower()
    label_blob = " ".join(labels).lower()
    for ttype, kws in _TYPE_RULES:
        if any(k in blob for k in kws):
            return ttype
    # Label-driven fallbacks for untitled tables.
    if "unlevered" in label_blob or "levered cf" in label_blob or "net operating income" in label_blob:
        return "cashflow_rollup"
    if "interest expense" in label_blob and "amortization" in label_blob:
        return "debt_schedule"
    return f"{periodicity}_table"


def _find_title(grid: list[tuple], header_row: int, label_col: int) -> str | None:
    """Nearest non-empty label-column text above the header (the table caption)."""
    for r in range(header_row, max(0, header_row - 6), -1):
        for c in (label_col, 1, 2):
            v = _gcell(grid, r, c)
            if isinstance(v, str) and v.strip() and _as_date(v) is None:
                # skip the header row's own period labels ("Total", a date, etc.)
                if r == header_row and c >= label_col:
                    continue
                return v.strip()
    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_workbook_tables(file_path: str | Path) -> list[dict]:
    """Detect and structure every period-axis table in the workbook."""
    file_path = Path(file_path)
    # read_only here is FAST because we pull each sheet's grid once with
    # iter_rows(values_only=True) (sequential) and then index in memory — random
    # ws.cell() access in read_only mode would be pathologically slow.
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    tables: list[dict] = []

    for sheet in wb.sheetnames:
        if sheet.rstrip(">").strip() == "" or sheet.endswith(">"):
            continue  # divider sheets ("PM>>>>", "MODEL>>>")
        ws = wb[sheet]
        if not ws.max_row or ws.max_row > _MAX_SHEET_ROWS:
            continue

        # Pull the scan window into memory once (sequential = fast).
        grid = [
            row for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row, _MAX_SCAN_ROWS),
                max_col=min(ws.max_column or 1, _MAX_SCAN_COLS), values_only=True,
            )
        ]

        claimed: set[int] = set()  # rows already inside a detected table
        for r in range(1, len(grid) + 1):
            if r in claimed:
                continue
            header = _detect_header(grid, r)
            if not header:
                continue
            label_col = _label_col(grid, r, header["period_cols"][0])
            body = _extract_rows(grid, r, label_col, header["period_cols"],
                                 header["date_headers"])
            if not body:
                continue
            title = _find_title(grid, r, label_col)
            ttype = _classify(title, [b["label"] for b in body], header["periodicity"])
            tables.append({
                "sheet":        sheet,
                "title":        title,
                "table_type":   ttype,
                "periodicity":  header["periodicity"],
                "period_axis":  "horizontal",
                "header_row":   r,
                "label_col":    label_col,
                "date_headers": header["date_headers"],
                "period_cols":  header["period_cols"],
                "rows":         body,
            })
            for b in body:
                claimed.add(b["row"])

    try:
        wb.close()
    except Exception:
        pass

    log.info("PARSE %s — %d table(s): %s", file_path.name, len(tables),
             ", ".join(f"{t['sheet']}:{t['table_type']}/{t['periodicity']}"
                       for t in tables[:8]))
    return tables


# Module-level parse cache (keyed by path + mtime). Persists across Streamlit
# reruns in the same process, so confirm / snapshot / deep-dives reuse one parse.
_PARSE_CACHE: dict[tuple[str, float], list[dict]] = {}


def parse_workbook_tables_cached(file_path: str | Path) -> list[dict]:
    p = Path(file_path)
    try:
        key = (str(p), p.stat().st_mtime)
    except OSError:
        return parse_workbook_tables(p)
    if key not in _PARSE_CACHE:
        _PARSE_CACHE[key] = parse_workbook_tables(p)
    return _PARSE_CACHE[key]


# Named cash-flow table types (analytically richest) vs generic periodic tables.
# build_time_series includes BOTH — consumers filter rows by keyword — but named
# types are emitted FIRST so they aren't crowded out of the row cap by input tabs.
_TS_NAMED = ("cashflow_rollup", "noi_schedule", "debt_schedule", "capex_schedule")
_TS_GENERIC = ("monthly_table", "quarterly_table", "semiannual_table",
               "annual_table", "weekly_table")
_TS_RELEVANT = set(_TS_NAMED) | set(_TS_GENERIC)


def _annualize_series(date_headers: list[str], values: list,
                      periodicity: str) -> tuple[list[str], list, str | None]:
    """Sum a sub-annual row into calendar-year totals (authoritative periodicity)."""
    if periodicity not in ("monthly", "quarterly", "semiannual"):
        return [], [], None
    annual: dict[str, float] = {}
    order: list[str] = []
    for h, v in zip(date_headers, values):
        if not isinstance(v, (int, float)):
            continue
        year = h[:4] if isinstance(h, str) and len(h) >= 4 else None
        if year is None:
            continue
        if year not in annual:
            annual[year] = 0.0
            order.append(year)
        annual[year] += float(v)
    if not annual:
        return [], [], None
    return order, [annual[y] for y in order], f"sum_{periodicity}_by_year"


def tables_to_time_series(tables: list[dict], max_rows: int = 80) -> list[dict]:
    """
    Convert parsed tables into time-series rows for the narrative layer, with the
    table's periodicity as the authoritative source (no per-row guessing) and
    calendar-year annualization for sub-annual tables.
    """
    out: list[dict] = []
    # Named cash-flow tables first so input-tab rows don't crowd out the cap.
    ordered = sorted(tables, key=lambda t: 0 if t["table_type"] in _TS_NAMED else 1)
    for t in ordered:
        if t["table_type"] not in _TS_RELEVANT:
            continue
        period = t["periodicity"]
        headers = t["date_headers"]
        label_col = t.get("label_col", 1)
        for row in t["rows"]:
            vals = row["values"]
            if not any(isinstance(v, (int, float)) and v != 0 for v in vals):
                continue
            ah, av, aggm = _annualize_series(headers, vals, period)
            out.append({
                "sheet":              t["sheet"],
                "label":              row["label"],
                "label_cell":         f"{get_column_letter(label_col)}{row['row']}",
                "headers":            headers[:12],
                "values":             vals[:12],
                "periodicity":        period,
                "annualized":         bool(ah),
                "aggregation_method": aggm,
                "annual_headers":     ah,
                "annual_values":      av,
                "table_type":         t["table_type"],
            })
            if len(out) >= max_rows:
                return out
    return out


def build_time_series(file_path: str | Path) -> list[dict]:
    """Parser-backed time series (authoritative periodicity). [] if no relevant tables."""
    return tables_to_time_series(parse_workbook_tables_cached(file_path))


def tag_metric_periodicity(tables: list[dict], records: dict[str, dict]) -> int:
    """
    Tag each metric record ({name: rec}) with the periodicity of the table its
    source cell belongs to. FLOW metrics only — ratio/percent (IRR, multiple,
    cap rate, LTV) are point-in-time even inside a periodic column, so tagging
    them would mislead. Mutates in place; returns count tagged.
    """
    tagged = 0
    for rec in records.values():
        if rec.get("unit") in ("ratio", "percent"):
            continue
        sheet, cell = rec.get("source_sheet"), rec.get("source_cell")
        if not sheet or not cell:
            continue
        t = table_for_cell(tables, sheet, cell)
        if t:
            rec["table_periodicity"] = t["periodicity"]
            rec["table_type"] = t["table_type"]          # #4: table membership
            if t.get("title"):
                rec["table_title"] = t["title"]
            tagged += 1
    return tagged


def table_for_cell(tables: list[dict], sheet: str, cell: str) -> dict | None:
    """Return the table whose period-axis body owns `sheet!cell`, if any."""
    from openpyxl.utils.cell import coordinate_to_tuple
    try:
        row, col = coordinate_to_tuple(cell)
    except Exception:
        return None
    for t in tables:
        if t["sheet"] != sheet:
            continue
        if col not in t["period_cols"]:
            continue
        if any(b["row"] == row for b in t["rows"]):
            return t
    return None


def periodicity_for_cell(tables: list[dict], sheet: str, cell: str) -> str | None:
    """
    Return the periodicity of the table that owns `sheet!cell`, if any. This is
    the integration hook: a metric extracted from a table cell INHERITS the
    table's periodicity instead of guessing per-row.
    """
    t = table_for_cell(tables, sheet, cell)
    return t["periodicity"] if t else None
